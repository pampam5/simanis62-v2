"""
Reports API endpoints untuk SIMANIS62 V2.

Menyediakan endpoints untuk export laporan KIB B dengan format 18 kolom BPAD DKI Jakarta.
"""

import io
import logging
from datetime import datetime
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.auth import CurrentUser, ExportUser
from app.core.database import get_db
from app.repositories.kib_repository import KibRepository
from app.schemas.response import PaginatedResponse, SuccessResponse

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/reports", tags=["Reports"])


@router.get(
    "/kib/b",
    response_model=PaginatedResponse[dict[str, Any]],
    summary="Get KIB B data",
    description="Ambil data KIB B dengan format 18 kolom BPAD DKI Jakarta.",
)
async def get_kib_b_data(
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
    ruangan_id: str | None = Query(None, description="Filter berdasarkan ruangan"),
    tahun_perolehan: int | None = Query(None, description="Filter berdasarkan tahun"),
    kondisi: str | None = Query(
        None, description="Filter berdasarkan kondisi (B/KB/RB)"
    ),
    page: int = Query(1, ge=1, description="Nomor halaman"),
    page_size: int = Query(20, ge=1, le=100, description="Jumlah item per halaman"),
) -> PaginatedResponse[dict[str, Any]]:
    """Get KIB B data dengan pagination.

    Args:
        ruangan_id: Filter berdasarkan ruangan
        tahun_perolehan: Filter berdasarkan tahun
        kondisi: Filter berdasarkan kondisi
        page: Nomor halaman
        page_size: Jumlah item per halaman

    Returns:
        PaginatedResponse dengan data KIB B
    """
    repo = KibRepository(db)

    # Get data
    skip = (page - 1) * page_size
    data = await repo.get_kib_b_data(
        ruangan_id=ruangan_id,
        tahun_perolehan=tahun_perolehan,
        kondisi=kondisi,
        skip=skip,
        limit=page_size,
    )

    # Get total count
    total = await repo.count_kib_b(
        ruangan_id=ruangan_id,
        tahun_perolehan=tahun_perolehan,
        kondisi=kondisi,
    )

    total_pages = (total + page_size - 1) // page_size

    return PaginatedResponse(
        data=data,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages,
    )


@router.get(
    "/kib/b/metadata",
    response_model=SuccessResponse[dict[str, Any]],
    summary="Get KIB B export metadata",
    description="Ambil metadata untuk preview sebelum export.",
)
async def get_kib_b_metadata(
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
    ruangan_id: str | None = Query(None, description="Filter berdasarkan ruangan"),
    tahun_perolehan: int | None = Query(None, description="Filter berdasarkan tahun"),
) -> SuccessResponse[dict[str, Any]]:
    """Get metadata untuk export KIB B.

    Args:
        ruangan_id: Filter berdasarkan ruangan
        tahun_perolehan: Filter berdasarkan tahun

    Returns:
        SuccessResponse dengan metadata export
    """
    repo = KibRepository(db)

    total_rows = await repo.count_kib_b(
        ruangan_id=ruangan_id,
        tahun_perolehan=tahun_perolehan,
    )

    total_nilai = await repo.get_total_nilai_kib_b(
        ruangan_id=ruangan_id,
        tahun_perolehan=tahun_perolehan,
    )

    metadata = {
        "provinsi": "DKI JAKARTA",
        "unit_organisasi": "DINAS PENDIDIKAN",
        "sub_unit_organisasi": "SDN 62 JAKARTA",
        "tanggal_export": datetime.now().strftime("%d/%m/%Y"),
        "total_rows": total_rows,
        "total_nilai": total_nilai,
    }

    return SuccessResponse(
        data=metadata,
        message="Metadata KIB B berhasil diambil",
    )


@router.get(
    "/export/kib-b",
    summary="Export KIB B to Excel",
    description="Export laporan KIB B ke Excel dengan format 18 kolom BPAD DKI Jakarta. Hanya Admin atau Viewer dengan dapat_ekspor=true.",
)
async def export_kib_b(
    current_user: ExportUser,
    db: AsyncSession = Depends(get_db),
    ruangan_id: str | None = Query(None, description="Filter berdasarkan ruangan"),
    tahun_perolehan: int | None = Query(None, description="Filter berdasarkan tahun"),
) -> StreamingResponse:
    """Export KIB B ke Excel.

    Format: 18 kolom BPAD DKI Jakarta
    - Harga dalam Rupiah penuh (bukan ribuan)
    - Tanggal dalam format DD/MM/YYYY

    Args:
        ruangan_id: Filter berdasarkan ruangan
        tahun_perolehan: Filter berdasarkan tahun

    Returns:
        StreamingResponse dengan file Excel

    Raises:
        HTTPException 403: Jika user tidak memiliki izin export
    """
    repo = KibRepository(db)

    # Get export data
    data = await repo.get_kib_b_export_data(
        ruangan_id=ruangan_id,
        tahun_perolehan=tahun_perolehan,
    )

    # Get total nilai
    total_nilai = await repo.get_total_nilai_kib_b(
        ruangan_id=ruangan_id,
        tahun_perolehan=tahun_perolehan,
    )

    # Generate Excel file
    try:
        excel_buffer = await _generate_kib_b_excel(data, total_nilai)
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error_code": "EXPORT_FAILED",
                "message": "Library openpyxl tidak tersedia. Silakan install dengan: pip install openpyxl",
            },
        )

    # Generate filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"KIB_B_{timestamp}.xlsx"

    logger.info(f"KIB B exported: {len(data)} items by user {current_user.id}")

    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get(
    "/can-export",
    response_model=SuccessResponse[dict[str, bool]],
    summary="Check export permission",
    description="Cek apakah user memiliki izin untuk export.",
)
async def can_export(
    current_user: CurrentUser,
) -> SuccessResponse[dict[str, bool]]:
    """Check apakah user bisa export.

    Admin: selalu bisa
    Viewer: hanya jika dapat_ekspor=true

    Returns:
        SuccessResponse dengan can_export boolean
    """
    can_export = current_user.role.value == "Admin" or current_user.dapat_ekspor

    return SuccessResponse(
        data={"can_export": can_export},
        message="Izin export berhasil dicek",
    )


async def _generate_kib_b_excel(
    data: list[dict[str, Any]], total_nilai: int
) -> io.BytesIO:
    """Generate Excel file untuk KIB B.

    Format: 18 kolom BPAD DKI Jakarta

    Args:
        data: List of export rows
        total_nilai: Total nilai untuk footer

    Returns:
        BytesIO buffer dengan Excel file
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "KIB B"

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    currency_format = "#,##0"

    # Title rows
    ws.merge_cells("A1:R1")
    ws["A1"] = "KARTU INVENTARIS BARANG (KIB) B"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:R2")
    ws["A2"] = "PERALATAN DAN MESIN"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:R3")
    ws["A3"] = f"Per Tanggal: {datetime.now().strftime('%d/%m/%Y')}"
    ws["A3"].alignment = Alignment(horizontal="center")

    # Empty row
    ws.merge_cells("A4:R4")

    # Header row (18 kolom BPAD DKI Jakarta)
    headers = [
        "NO",
        "KODE BARANG",
        "NAMA BARANG",
        "NO. REGISTER",
        "UKU-RAN",
        "SATU-AN",
        "TAHUN PEROLEHAN",
        "BA-HAN",
        "MEREK",
        "TYPE",
        "TGL. BPKB/TGL. DOK.",
        "NO. CHASIS/NO. RANGKA",
        "NO. MESIN/NO. PABRIK",
        "NOMOR POLISI",
        "KONDISI (B/KB/RB)",
        "HARGA (Rp.)",
        "KAPITALISASI (Rp.)",
        "TOTAL (Rp.)",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, row_data in enumerate(data, 6):
        values = [
            row_data["no"],
            row_data["kode_barang"],
            row_data["nama_barang"],
            row_data["nomor_register"],
            row_data["ukuran_cc"],
            row_data["satuan"],
            row_data["tahun_perolehan"],
            row_data["bahan"],
            row_data["merk"],
            row_data["tipe"],
            row_data["tanggal_dokumen"],
            row_data["nomor_rangka"],
            row_data["nomor_mesin"],
            row_data["nomor_polisi"],
            row_data["kondisi"],
            row_data["harga"],
            row_data["kapitalisasi"],
            row_data["total_harga"],
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # Format currency columns (16, 17, 18)
            if col_idx in [16, 17, 18]:
                cell.number_format = currency_format
                cell.alignment = Alignment(horizontal="right")

    # Footer row with totals
    footer_row = len(data) + 6
    ws.cell(row=footer_row, column=1, value="TOTAL")
    ws.cell(row=footer_row, column=1).font = Font(bold=True)
    ws.merge_cells(f"A{footer_row}:O{footer_row}")

    # Total harga
    total_harga = sum(row["harga"] for row in data)
    total_kapitalisasi = sum(row["kapitalisasi"] for row in data)
    total_total = sum(row["total_harga"] for row in data)

    ws.cell(row=footer_row, column=16, value=total_harga)
    ws.cell(row=footer_row, column=16).number_format = currency_format
    ws.cell(row=footer_row, column=16).font = Font(bold=True)

    ws.cell(row=footer_row, column=17, value=total_kapitalisasi)
    ws.cell(row=footer_row, column=17).number_format = currency_format
    ws.cell(row=footer_row, column=17).font = Font(bold=True)

    ws.cell(row=footer_row, column=18, value=total_total)
    ws.cell(row=footer_row, column=18).number_format = currency_format
    ws.cell(row=footer_row, column=18).font = Font(bold=True)

    # Auto-adjust column widths
    column_widths = [5, 15, 30, 10, 10, 8, 8, 12, 15, 15, 12, 20, 20, 12, 8, 15, 15, 15]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
