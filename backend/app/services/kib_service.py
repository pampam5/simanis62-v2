"""KIB service untuk SIMANIS62 V2.

Module ini menyediakan KibService untuk:
- Generate laporan KIB A-F
- Export ke Excel dengan format BPAD DKI Jakarta
"""

from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.models.aset import Aset, KategoriKIB, Kondisi
from app.repositories.aset_repository import AsetRepository
from app.schemas.kib import (
    KibBItem,
    KibExportRequest,
    KibExportResponse,
    KibItemBase,
    KibReportResponse,
    KibSummary,
)
from app.services.base import BaseService

# KIB Category Names
KIB_NAMES = {
    KategoriKIB.A: "Tanah",
    KategoriKIB.B: "Peralatan dan Mesin",
    KategoriKIB.C: "Gedung dan Bangunan",
    KategoriKIB.D: "Jalan, Irigasi, dan Jaringan",
    KategoriKIB.E: "Aset Tetap Lainnya",
    KategoriKIB.F: "Konstruksi dalam Pengerjaan",
}

# KIB B Columns (18 kolom BPAD DKI Jakarta)
KIB_B_COLUMNS = [
    "No",
    "Kode Barang",
    "Nama Barang",
    "Nomor Register",
    "Satuan",
    "Ukuran/CC",
    "Bahan",
    "Merk/Type",
    "Tahun Perolehan",
    "Nomor Rangka",
    "Nomor Mesin",
    "Nomor Polisi",
    "Tanggal Dokumen",
    "Kondisi",
    "Asal Usul",
    "Harga",
    "Kapitalisasi",
    "Keterangan",
]


class KibService(BaseService[Aset, AsetRepository]):
    """Service untuk generate laporan KIB dan export Excel.

    Menyediakan:
    - get_kib_report: Generate laporan KIB untuk kategori tertentu
    - export_to_excel: Export laporan ke Excel

    Example:
        ```python
        service = KibService(session)
        report = await service.get_kib_report(KategoriKIB.B)
        export = await service.export_to_excel(request)
        ```
    """

    def __init__(self, session: AsyncSession) -> None:
        """Initialize KibService.

        Args:
            session: AsyncSession untuk database operations.
        """
        super().__init__(session, AsetRepository(session), "KibService")

    async def get_kib_report(
        self,
        kategori_kib: KategoriKIB,
        tahun: int | None = None,
        ruangan_id: str | None = None,
        page: int = 1,
        page_size: int = 1000,
    ) -> KibReportResponse:
        """Generate laporan KIB untuk kategori tertentu.

        Args:
            kategori_kib: Kategori KIB (A-F).
            tahun: Filter berdasarkan tahun perolehan (optional).
            ruangan_id: Filter berdasarkan ruangan (optional).
            page: Nomor halaman.
            page_size: Jumlah item per halaman.

        Returns:
            KibReportResponse dengan summary dan items.
        """
        self.log_info(f"Generating KIB {kategori_kib.value} report")

        skip = (page - 1) * page_size

        # Get assets for report
        assets = await self.repository.get_for_kib_report(
            kategori_kib=kategori_kib,
            tahun=tahun,
            ruangan_id=ruangan_id,
            skip=skip,
            limit=page_size,
        )

        # Get counts and totals
        total_count = await self.repository.count_for_kib_report(
            kategori_kib=kategori_kib,
            tahun=tahun,
            ruangan_id=ruangan_id,
        )

        total_value = await self.repository.get_total_value_for_kib(
            kategori_kib=kategori_kib,
            tahun=tahun,
        )

        # Count by kondisi
        kondisi_counts = self._count_by_kondisi(assets)

        # Build summary
        summary = KibSummary(
            kategori_kib=kategori_kib,
            total_item=total_count,
            total_nilai=total_value,
            kondisi_baik=kondisi_counts.get(Kondisi.BAIK, 0),
            kondisi_rusak_ringan=kondisi_counts.get(Kondisi.RUSAK_RINGAN, 0),
            kondisi_rusak_berat=kondisi_counts.get(Kondisi.RUSAK_BERAT, 0),
        )

        # Convert to KIB items
        items = [self._to_kib_item(a, i + 1 + skip) for i, a in enumerate(assets)]

        self.log_info(f"KIB {kategori_kib.value} report generated: {total_count} items")

        return KibReportResponse(
            kategori_kib=kategori_kib,
            nama_sekolah=settings.nama_sekolah,
            tanggal_cetak=datetime.now(UTC),
            summary=summary,
            items=items,
        )

    async def export_to_excel(self, request: KibExportRequest) -> KibExportResponse:
        """Export laporan KIB ke Excel.

        Args:
            request: KibExportRequest dengan parameter export.

        Returns:
            KibExportResponse dengan info file yang di-generate.
        """
        self.log_info(f"Exporting KIB {request.kategori_kib.value} to Excel")

        # Get all assets for export (no pagination)
        assets = await self.repository.get_for_kib_report(
            kategori_kib=request.kategori_kib,
            tahun=request.tahun_perolehan,
            ruangan_id=request.ruangan_id,
            skip=0,
            limit=10000,  # Max export
        )

        # Filter by kondisi if needed
        if not request.include_rusak:
            assets = [a for a in assets if a.kondisi == Kondisi.BAIK]

        # Generate filename
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"KIB_{request.kategori_kib.value}_{timestamp}.xlsx"

        # Create export directory
        export_dir = Path(settings.export_dir)
        export_dir.mkdir(parents=True, exist_ok=True)
        file_path = export_dir / filename

        # Generate Excel file
        file_size = await self._generate_excel(assets, request.kategori_kib, file_path)

        self.log_info(f"Excel exported: {filename}, {len(assets)} items")

        return KibExportResponse(
            filename=filename,
            file_path=str(file_path),
            file_size=file_size,
            total_items=len(assets),
            kategori_kib=request.kategori_kib,
        )

    async def _generate_excel(
        self,
        assets: list[Aset],
        kategori_kib: KategoriKIB,
        file_path: Path,
    ) -> int:
        """Generate Excel file untuk KIB report.

        Args:
            assets: List of assets to export.
            kategori_kib: Kategori KIB.
            file_path: Path untuk output file.

        Returns:
            File size in bytes.
        """
        try:
            # Try to use openpyxl (ClosedXML equivalent for Python)
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = f"KIB {kategori_kib.value}"

            # Styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Title row
            ws.merge_cells("A1:R1")
            ws["A1"] = f"KARTU INVENTARIS BARANG (KIB) {kategori_kib.value}"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = Alignment(horizontal="center")

            # School name row
            ws.merge_cells("A2:R2")
            ws["A2"] = settings.nama_sekolah
            ws["A2"].font = Font(bold=True, size=12)
            ws["A2"].alignment = Alignment(horizontal="center")

            # Empty row
            ws.merge_cells("A3:R3")

            # Header row
            columns = self._get_columns_for_kib(kategori_kib)
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=4, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Data rows
            for row_idx, aset in enumerate(assets, 5):
                row_data = self._get_row_data(aset, row_idx - 4, kategori_kib)
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    if col_idx == len(row_data) - 2:  # Harga column
                        cell.number_format = "#,##0"

            # Footer row with totals
            footer_row = len(assets) + 5
            ws.cell(row=footer_row, column=1, value="TOTAL")
            ws.cell(row=footer_row, column=1).font = Font(bold=True)

            total_harga = sum(a.harga for a in assets)
            harga_col = len(columns) - 2  # Harga is usually second to last
            ws.cell(row=footer_row, column=harga_col, value=total_harga)
            ws.cell(row=footer_row, column=harga_col).number_format = "#,##0"
            ws.cell(row=footer_row, column=harga_col).font = Font(bold=True)

            # Auto-adjust column widths
            for col_idx in range(1, len(columns) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15

            # Save file
            wb.save(file_path)

            return file_path.stat().st_size

        except ImportError:
            # Fallback: create simple CSV if openpyxl not available
            self.log_warning("openpyxl not available, creating CSV instead")
            csv_path = file_path.with_suffix(".csv")
            return await self._generate_csv(assets, kategori_kib, csv_path)

    async def _generate_csv(
        self,
        assets: list[Aset],
        kategori_kib: KategoriKIB,
        file_path: Path,
    ) -> int:
        """Generate CSV file as fallback.

        Args:
            assets: List of assets to export.
            kategori_kib: Kategori KIB.
            file_path: Path untuk output file.

        Returns:
            File size in bytes.
        """
        import csv

        columns = self._get_columns_for_kib(kategori_kib)

        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(columns)

            for idx, aset in enumerate(assets, 1):
                row_data = self._get_row_data(aset, idx, kategori_kib)
                writer.writerow(row_data)

        return file_path.stat().st_size

    def _get_columns_for_kib(self, kategori_kib: KategoriKIB) -> list[str]:
        """Get column headers for KIB category.

        Args:
            kategori_kib: Kategori KIB.

        Returns:
            List of column headers.
        """
        if kategori_kib == KategoriKIB.B:
            return KIB_B_COLUMNS
        # Generic columns for other KIB types
        return [
            "No",
            "Kode Barang",
            "Nama Barang",
            "Nomor Register",
            "Tahun Perolehan",
            "Kondisi",
            "Asal Usul",
            "Harga",
            "Keterangan",
        ]

    def _get_row_data(
        self, aset: Aset, row_num: int, kategori_kib: KategoriKIB
    ) -> list[Any]:
        """Get row data for asset.

        Args:
            aset: Aset model.
            row_num: Row number.
            kategori_kib: Kategori KIB.

        Returns:
            List of cell values.
        """
        kondisi_map = {
            Kondisi.BAIK: "B",
            Kondisi.RUSAK_RINGAN: "KB",
            Kondisi.RUSAK_BERAT: "RB",
        }

        if kategori_kib == KategoriKIB.B:
            return [
                row_num,
                aset.kode_barang,
                aset.nama_barang,
                aset.nomor_register,
                getattr(aset, "satuan", ""),
                getattr(aset, "ukuran_cc", ""),
                getattr(aset, "bahan", ""),
                f"{getattr(aset, 'merk', '')} {getattr(aset, 'tipe', '')}".strip(),
                aset.tahun_perolehan,
                getattr(aset, "nomor_rangka", ""),
                getattr(aset, "nomor_mesin", ""),
                getattr(aset, "nomor_polisi", ""),
                getattr(aset, "tanggal_dokumen", ""),
                kondisi_map.get(aset.kondisi, ""),
                aset.asal_usul.value if aset.asal_usul else "",
                aset.harga,
                getattr(aset, "kapitalisasi", ""),
                aset.keterangan or "",
            ]
        return [
            row_num,
            aset.kode_barang,
            aset.nama_barang,
            aset.nomor_register,
            aset.tahun_perolehan,
            kondisi_map.get(aset.kondisi, ""),
            aset.asal_usul.value if aset.asal_usul else "",
            aset.harga,
            aset.keterangan or "",
        ]

    def _count_by_kondisi(self, assets: list[Aset]) -> dict[Kondisi, int]:
        """Count assets by kondisi.

        Args:
            assets: List of assets.

        Returns:
            Dict with kondisi counts.
        """
        counts: dict[Kondisi, int] = {}
        for aset in assets:
            counts[aset.kondisi] = counts.get(aset.kondisi, 0) + 1
        return counts

    def _to_kib_item(self, aset: Aset, nomor_urut: int) -> KibItemBase:
        """Convert Aset to KIB item schema.

        Args:
            aset: Aset model.
            nomor_urut: Nomor urut dalam laporan.

        Returns:
            KibItemBase or specific KIB item schema.
        """
        kondisi_map = {
            Kondisi.BAIK: "B",
            Kondisi.RUSAK_RINGAN: "KB",
            Kondisi.RUSAK_BERAT: "RB",
        }

        base_data = {
            "nomor_urut": nomor_urut,
            "kode_barang": aset.kode_barang,
            "nama_barang": aset.nama_barang,
            "nomor_register": aset.nomor_register,
            "kondisi": kondisi_map.get(aset.kondisi, ""),
            "tahun_perolehan": aset.tahun_perolehan,
            "asal_usul": aset.asal_usul.value if aset.asal_usul else "",
            "harga": aset.harga,
            "keterangan": aset.keterangan,
        }

        if aset.kategori_kib == KategoriKIB.B:
            return KibBItem(
                **base_data,
                satuan=getattr(aset, "satuan", None),
                ukuran_cc=getattr(aset, "ukuran_cc", None),
                bahan=getattr(aset, "bahan", None),
                merk=getattr(aset, "merk", None),
                tipe=getattr(aset, "tipe", None),
                nomor_rangka=getattr(aset, "nomor_rangka", None),
                nomor_mesin=getattr(aset, "nomor_mesin", None),
                nomor_polisi=getattr(aset, "nomor_polisi", None),
                tanggal_dokumen=getattr(aset, "tanggal_dokumen", None),
                kapitalisasi=getattr(aset, "kapitalisasi", None),
                total_harga=getattr(aset, "total_harga", None),
            )

        return KibItemBase(**base_data)
